import os
from copy import copy
import datetime
import subprocess
import shutil
import openpyxl
from openpyxl.styles import PatternFill
import re
import time
from PyQt5.QtWidgets import *
from PyQt5 import QtTest
import operator

from Utilities import Utilities
from Logger import *
from Exceptions import TestError, ParameterError, SimEngineInjectionError, GDTConnectionError, SimEngineConnectionError
from GDTInterface import GDTInterface
from PIL import Image
from PCSIMSocketInterface import PCSIMSocketInterface
from SVNInterface import SVNInterface
from SimEngineInterface import SimEngineInterface
from ImageComparator import ImageComparator, ImageComparatorError
from VideoComparator import VideoComparator

# Global Defines
simulation = True
SLEEP_TIME = 100  # default sleep time im ms


class TestClass(object):
    def __init__(self, _test_name, _svn_download, _svn_results, _executed, _host_env=False, _menu=None, _ci=False,
                 _instrumented=False, _called=False):
        super(TestClass, self).__init__()

        self.test_name = _test_name
        self.test_path = Utilities.get_tests_folder()

        self.test_path = f'{self.test_path}/{self.test_name}'
        self.output_path = f'{self.test_path}/Output'
        self.result_path = f'{self.test_path}/Results'
        self.golden_images_path = f'{self.test_path}/GoldenImage'
        self.excel_file = f'{self.test_path}/{self.test_name}.xlsx'
        self.excel_result_file = f'{self.test_path}/{self.test_name}_Result.xlsx'
        self.host_env = _host_env
        self.executed = _executed
        self.menu = _menu
        self.ci = _ci
        self.called = _called

        self.svn_interface = SVNInterface(self.test_name, self.ci)

        self.SVNDownload = _svn_download
        self.SVNResults = _svn_results
        if self.SVNResults:
            if self.svn_interface.result_path.split('/')[0] == "":
                if self.ci:
                    svn_result_path = "CIPATH"
                else:
                    svn_result_path = "RESULTPATH"
                raise TestError(self.test_name,
                                f"There is no definition for \"{svn_result_path}\" in ConnectionsConfig.xml"
                                f"\nCannot use svn_commit!!")
            if self.svn_interface.user == "N/A" or self.svn_interface.user == "":
                raise TestError(self.test_name,
                                f"SVN User is not defined in ConnectionsConfig.xml\nCannot use svn_commit!!")
            if self.svn_interface.password == "N/A" or self.svn_interface.password == "":
                raise TestError(self.test_name,
                                f"SVN Password is not defined in ConnectionsConfig.xml\nCannot use svn_commit!!")

        if self.SVNDownload is False:
            if not os.path.exists(self.test_path):
                raise TestError(self.test_name,
                                f"There are no tests corresponding to the specified test name parameter: {self.test_name}")

            if not os.path.exists(self.excel_file):
                raise TestError(self.test_name, "There is no excel file inside the test folder "
                                               "or excel file does not match test folder name.")

        if self.SVNDownload is True:
            if self.svn_interface.exists() is False:
                raise TestError(self.test_name,
                                f"There are no tests corresponding to the specified test name parameter: {self.test_name}")

        if self.host_env is True:
            self.socket_interface = PCSIMSocketInterface()

        self.GDTInterfaces = [GDTInterface(interface) for interface in Utilities.get_gdt_interface_connection()]
        try:
            self.SimEngineInterface = SimEngineInterface()
        except SimEngineConnectionError as error:
            raise TestError(self.test_name, error.message)

        self.error_log_file = f'{self.test_path}/Error.txt'
        self.normal_log_file = f'{self.test_path}/Log.txt'
        self.video_compression_log_file = f'{self.test_path}/VideoCompression.txt'

        self.currentsheet = ''
        self.scenario_id = 0
        self.ws = None
        self.wb = None
        self.passed = False
        self.caller = None
        self.startfrom = 0
        self.current_status = "N/A"

        self.manualRun = False

        self.gdt_injections = [dict(), dict()]
        self.sim_injections = dict()

        self.normal_logger = None
        self.error_logger = None

        self.actions = dict()
        self.video_end_time = None
        self.instrumented = _instrumented
        self.instrumented_port = None
        if self.instrumented:
            self.instrumented_port = subprocess.Popen(f"./lib/PortCapture.exe 7100 \"{self.test_path}/{self.test_name}.exh\"")

    def consoleprint(self, text, color=''):
        if not self.executed:
            self.menu.onUpdateText(text, color)
        # if(self.executed == True):
        else:
            print(text)

    @staticmethod
    def sleep(milliseconds):
        QtTest.QTest.qWait(milliseconds)

    def run_test(self):
        try:
            if self.called and self.menu is not None:
                self.menu.updateScenario(f"Test Call: {self.test_name}", -1)

            # If Downloading from svn, ask to remove old folder
            # If was executed from cmd automatically override test data
            if self.SVNDownload:
                if self.svn_interface.exists() is False:
                    raise TestError(self.test_name, "Could not find test folder in the SVN.")
                if os.path.exists(self.test_path):
                    self.consoleprint("Overriding test data..")
                    shutil.rmtree(self.test_path)
                    self.sleep(1000)
                self.svn_interface.export(self.test_path)

            # Logging file
            if os.path.isfile(self.normal_log_file):
                os.remove(self.normal_log_file)
            self.normal_logger = Logger("Normal Log", self.normal_log_file)

            # Error logging
            if os.path.isfile(self.error_log_file):
                os.remove(self.error_log_file)
            self.error_logger = Logger("Error Log", self.error_log_file)

            # Error logging
            if os.path.isfile(self.video_compression_log_file):
                os.remove(self.video_compression_log_file)
            self.video_compression_log = open(self.video_compression_log_file, 'a')

            # Open the excel
            try:
                self.wb = openpyxl.load_workbook(self.excel_file, data_only=True)  # Open with readonly
            except FileNotFoundError:
                raise TestError(self.test_name, "There is no excel file inside the test folder "
                                               "or excel file does not match test folder name.")

            sheetnames = self.wb.sheetnames
            self.ws = self.wb.active

            # Receive name of sheets from the excel, look for 'scenarios' if there is more than 1 sheet
            if len(sheetnames) > 1:
                self.ws = self.wb[sheetnames[0]]
                try:
                    self.ws = self.wb["Scenarios"]
                except KeyError:
                    raise TestError(self.test_name, "Sheet 'Scenarios' was not found, please check your excel\n"
                                                   f"Sheet names: {sheetnames}")

            # Remove Output folder
            if os.path.exists(self.output_path):
                try:
                    shutil.rmtree(self.output_path)
                    self.sleep(1500)
                except PermissionError and OSError:
                    raise TestError(self.test_name, "Output folder is in use!\nPlease close any open files in it")
            os.mkdir(self.output_path)

            # Remove results folder
            if simulation:
                if os.path.exists(self.result_path):
                    try:
                        shutil.rmtree(self.result_path)
                        self.sleep(1000)
                    except PermissionError and OSError:
                        raise TestError(self.test_name, "Results folder is in use!\nPlease close any open files in it")
                os.mkdir(self.result_path)

            # Remove results file
            if os.path.isfile(self.excel_result_file):
                try:
                    os.remove(self.excel_result_file)
                except PermissionError:
                    raise TestError(self.test_name, "Results excel is in use!\nPlease close before trying to run.")
                self.sleep(1000)

            self.sleep(2000)

            # If manual run and local run
            if self.SVNResults and self.menu is not None:
                if not self.SVNDownload:
                    self.manualRun = True

            time_string = datetime.datetime.now()
            self.log_write("%" * 80)
            self.log_write("TEST: " + self.test_name)
            self.log_write("TIME&DATE: %s" % (time_string.strftime("%d/%m/%Y %X")))
            self.log_write(f"VERSION: {Utilities.get_current_version()}")
            self.log_write("%" * 80)

            # ws = wb.active
            if self.called:
                # If test is called copy worksheet to the caller's workbook
                ws = self.ws
                copy_ws = self.caller.wb[self.test_name]

                for idx, cd in ws.column_dimensions.items():
                    copy_ws.column_dimensions[idx].width = cd.width
                for idx, rd in ws.row_dimensions.items():
                    copy_ws.row_dimensions[idx].height = rd.height

                for row in ws.rows:
                    for cell in row:
                        new_cell = copy_ws.cell(row=cell.row, column=cell.col_idx, value=cell.value)
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = copy(cell.number_format)
                            new_cell.protection = copy(cell.protection)
                            new_cell.alignment = copy(cell.alignment)

                self.caller.wb.save(self.caller.excel_result_file)

            if "Actions" in sheetnames:
                self.consoleprint("\nSpecial Actions:")
                self.action_sheet(self.wb["Actions"])

            if "Preconditions" in sheetnames:
                self.consoleprint("\nPreconditions:")
                self.run_sheet(self.wb["Preconditions"], True)

            self.consoleprint("\nTest Scenarios:")
            self.ws = self.wb.active
            for sheet in sheetnames:
                if "Scenarios" in sheet or len(sheetnames) == 1:
                    self.consoleprint("\nTest Scenarios for sheet %s:" % sheet)

                    if len(sheetnames) > 1:
                        self.log_write("-" * 80)
                        self.log_write("SHEET: " + sheet)
                        self.log_write("-" * 80)
                    self.ws = self.wb[sheet]

                    # Create output folder for current sheet
                    output_folder = f"{self.output_path}/{sheet}"
                    os.mkdir(output_folder)
                    if simulation:
                        # Create results folder for current sheet
                        results_folder = f"{self.result_path}/{sheet}"
                        os.mkdir(results_folder)
                    self.currentsheet = sheet
                    if self.menu is not None:
                        self.menu.setWindowTitle(
                            f"{Utilities.get_current_version()} - Running Test: {self.test_name} - Sheet: {sheet}")
                    self.run_sheet(self.ws, False, self.startfrom)

        except KeyboardInterrupt:
            self.error_log("Test run cancelled by user.")
        except TestError as error:
            self.error_log(f"\nERROR WHILE RUNNING: {error.message}")
        except ParameterError as error:
            self.error_log(f"\nPARAMETER ERROR: {error.parameter}: {error.message}")
        except GDTConnectionError as error:
            self.error_log(f"\nGDT ERROR: {error.message}")
        except StopIteration:
            pass
        else:
            # If no exceptions set passed to true
            self.passed = True
        finally:
            self.test_end(self.passed)

    def test_end(self, passed):
        if self.instrumented_port is not None:
            if passed:
                self.gdt_inject("ifSCAStartDataDownloadOFP", 1, "GDT")
                self.instrumented_port.wait()
            else:
                self.consoleprint("Could not download instrumented history file due to test failure.")

        if self.menu is not None:
            self.menu.setWindowTitle(Utilities.get_current_version())
            self.menu.pauseButton.setEnabled(False)
            # Re-enable menu buttons
            self.menu.backButton.setEnabled(True)

        # Clear SIM & GDT Injections
        self.clear_injections()

        # Remove Logging
        del self.normal_logger
        del self.error_logger
        self.video_compression_log.close()

        # Remove socket connection for host_env
        if self.host_env:
            del self.socket_interface

        # Print AutoTester version in the excel
        self.ws.cell(row=self.ws.max_row + 2, column=1,
                     value=f'Test was executed with {Utilities.get_current_version()}')

        # Save results excel
        if not self.called:
            self.wb.save(self.excel_result_file)

        # Test end
        if passed:
            # If upload results to SVN was checked, upload results
            if self.SVNResults and self.SVNDownload and self.called is False:
                self.consoleprint("\nUploading results to svn..")
                date = time.strftime("%d_%m_%Y")
                self.svn_interface.upload_file(f"{self.test_path}/Log.txt", f"{date}/Log.txt")
                self.svn_interface.upload_file(f"{self.test_path}/Error.txt", f"{date}/Error.txt")
                self.svn_interface.upload_file(f"{self.test_path}/Results", f"{date}/Results")
                self.svn_interface.upload_file(f"{self.test_path}/Output", f"{date}/Output")
                self.svn_interface.upload_file(f"{self.excel_result_file}", f"{date}/{self.test_name}_Result.xlsx")

                # Remove tests folder
                Utilities.remove_folder(self.test_path)

                self.sleep(1500)
                self.consoleprint("\n..Done")

            if self.menu is not None and self.called is False:
                self.menu.popup(Utilities.get_current_version(), "Test run was done!", 1)

            if self.called:
                self.consoleprint("=" * 50)
                # self.menu.updateScenario(" ", 0)

            self.consoleprint("\n\n\nTest run is done.", "grey")

    def clear_injections(self):
        if len(self.sim_injections.items()) > 0:
            self.consoleprint("\nClearing SIM Injections")
            for label, value in self.sim_injections.items():
                try:
                    self.SimEngineInterface.inject_value(label, 'N/A', '')
                except SimEngineInjectionError:
                    pass
        self.sim_injections = dict()

        self.sleep(1200)

        self.SimEngineInterface.unapply_values_on_exit()
        self.SimEngineInterface.unapply_values()
        self.SimEngineInterface.set_noise(False)
        self.consoleprint("SIM: Noise has been disabled.", "#3CB371")

        if hasattr(self, 'GDTInterfaces'):
            for ofp, interface in enumerate(self.GDTInterfaces):
                if self.GDTInterfaces[ofp].is_connected:
                    self.consoleprint(f"\nClearing GDT Injections for {self.GDTInterfaces[ofp].connection}")
                    for key, value in self.gdt_injections[ofp].items():
                        # print(value[1], key, False)
                        self.GDTInterfaces[ofp].inject_data_item_override(value[1], key, False)
                    self.gdt_injections[ofp] = dict()
                    self.consoleprint(f"GDT: {self.GDTInterfaces[ofp].connection} disconnected successfully.", "#3CB371")
                    del self.GDTInterfaces[ofp]

        if self.instrumented_port is not None:
            self.instrumented_port.terminate()

    def action_sheet(self, ws):
        for row in range(1, int(ws.max_row) + 1):
            rows = ws.iter_rows(min_row=row, max_row=row)
            first_row = next(rows)
            line = [c.value for c in first_row]
            for data in line:
                if data is not None:
                    try:
                        shortcut, real_action = data.split('=')
                        self.actions[shortcut.lower()] = real_action
                        self.consoleprint(f"{shortcut} = {real_action}")
                    except AttributeError and IndexError:
                        pass
        # self.consoleprint(self.actions)

    def run_sheet(self, ws, preconditions=False, startfrom=0):
        """Input: Sheet Object, and is it a precondition sheet or not"""

        self.ws = ws
        hash_tag = self.find_hash_tag()
        if hash_tag == [0, 0]:
            self.consoleprint("Cannot find the symbol '#' to determinate where to read excel", "#FF0000")
            return None
        _startFromScenario = 0
        if startfrom != 0:
            _startFromScenario = self.find_scenario(startfrom, hash_tag[1])
            if _startFromScenario == 0 and startfrom != 0:
                self.consoleprint("Could not found the requested scenario, running from start.", "#FAFAFA")
        # print("scenario_find in: row: %d, col: %d"%(_startFromScenario,hash_tag[1]))
        if _startFromScenario == 0:
            start = hash_tag[0]
        else:
            start = _startFromScenario - 2
        scenario_col = hash_tag[1]

        # Create new columns
        if preconditions is False:
            column = self.ws.max_column + 1
            self.write_line(start, column, "Pass/Fail")
            self.write_line(start + 1, column, "Empty")
            self.write_line(start, column + 1, "Actual")
            self.write_line(start + 1, column + 1, "Empty")
            self.write_line(start, column + 2, "PR")
            self.write_line(start + 1, column + 2, "Empty")

        # Get the first 3 rows
        rows = ws.iter_rows(min_row=hash_tag[0] - 1, max_row=hash_tag[0] + 1)

        # define first row as reference (GDT/SIM/ACTION/EXPECTED)
        reference = [c.value for c in next(rows)]

        # define second row as dataitem/labels
        headings = [c.value for c in next(rows)]

        # define third row as free text
        free_text = [c.value for c in next(rows)]

        self.SimEngineInterface.set_noise(True)
        self.consoleprint("SIM: Noise has been enabled.", "#3CB371")

        for x in reference:
            if x is None:
                continue
            if x.upper() == "GDT":
                self.gdt_connect(0)
            if x.upper() == "GDT_PSP2":
                self.gdt_connect(1)

        # +(startfrom-1)
        if _startFromScenario != 0:
            self.scenario_id = 0 + (startfrom - 1)
        else:
            self.scenario_id = 0
        new_line = False
        if self.menu is not None:
            self.steps = (ws.max_row + 1) - start - 2
            self.progressSteps = 0
            self.menu.progressBar.setRange(0, self.steps)
            if preconditions:
                self.update_current_test_status("PRECONDITION")
            else:
                self.update_current_test_status("PASSED")
        try:
            # Start running on the execl
            for row, line in enumerate(ws.iter_rows(min_row=start+2, max_row=ws.max_row+1), start+2):
                if self.called:
                    wsx = self.caller.wb[self.test_name]
                else:
                    wsx = self.ws
                if preconditions is False:
                    self.write_line(row, wsx.max_column - 2, "Empty")  # PASS/FAIL
                    self.write_actual(row, wsx.max_column - 1, " ")  # ACTUAL
                    self.write_line(row, wsx.max_column, "Empty")  # PR
                for column, cell in enumerate(line):
                    value = cell.value
                    if self.menu is not None:
                        pause = self.menu.pauseButton.isChecked()
                        while pause is False:
                            pause = self.menu.pauseButton.isChecked()
                            self.sleep(1000)
                    reference_name = reference[column]
                    if value == 'N\A':
                        value = 'N/A'
                    if value is not None and reference_name is not None and line[scenario_col - 1].value is not None:
                        if value == 'N/A' and reference_name.upper() != 'EXPECTED':
                            continue
                        if preconditions is False:
                            if new_line is False:
                                self.scenario_id += 1
                                new_line = True
                                self.consoleprint("\nScenario: %d" % self.scenario_id, "#ff6f61")
                                self.log_write("")
                                self.log_write("\t" + "=" * 60)
                                self.log_write("\tScenario %d:" % self.scenario_id)
                                self.log_write("\t" + "=" * 60)
                        # Handle column by referring to header data
                        if re.findall("^GDT_STRUCT_PSP2|^GDT_STRUCT|^GDT_PSP2|^GDT|^SIM", reference_name.upper()):
                            if re.findall("INJECT*", headings[column]):
                                self.handle_inject(reference_name.upper(), value)
                            else:
                                if reference_name.upper() == "SIM":
                                    self.sim_inject(headings[column], value)
                                if reference_name.upper() == "GDT" or reference_name.upper() == "GDT_PSP2":
                                    self.gdt_inject(headings[column], value, reference_name.upper())
                        if reference_name.upper() == "ACTION":
                            if re.findall("^VIDEO|^BVIDEO", value.upper()):
                                self.write_actual(row, wsx.max_column - 1, "Video")

                            if re.findall("^CALL_TEST", value.upper()):
                                self.handle_action(value, row, wsx.max_column)
                            else:
                                self.handle_action(value)

                        if reference_name.upper() == "EXPECTED":
                            # Wait for background video to end before running expected column
                            if self.video_end_time is not None:
                                self.consoleprint("Waiting for video to end...")
                                while time.time() < self.video_end_time:
                                    self.sleep(1000)
                                self.consoleprint("...Done")
                                self.video_end_time = None

                                uncompressed_video_path = f'{self.result_path}/{self.currentsheet}/' \
                                                          f'Scenario_{self.scenario_id}.avi'
                                compressed_video_path = uncompressed_video_path.replace(".avi", ".mp4")
                                self.compress_video(uncompressed_video_path, compressed_video_path)
                            # Expected method
                            try:
                                if (' ' in value) or ('\n' in value) or ('\t' in value):
                                    raise TestError(self.test_name,
                                                    f'Expected commands shall not have whitespaces, new lines or tabulations: {value}')

                                # Split expected data by ',' to handle multiple data
                                expected_data = re.split(r',\s*(?![^()]*\))', str(value))
                                pass_fail = "PASSED"
                                result_string = ""
                                results_string_failed = ''
                                results_string_passed = ''
                                for i in expected_data:
                                    result = self.handle_expected(i)
                                    result_string += result[1]
                                    if result[0] == "FAILED":
                                        pass_fail = "FAILED"
                                        if len(result) > 3:
                                            if result[3] is not '':
                                                results_string_failed += result[3]
                                    if result[0] == "PASSED":
                                        if len(result) > 2:
                                            if result[2] is not '':
                                                results_string_passed += result[2]

                                result = [pass_fail, result_string, results_string_passed, results_string_failed]
                                if result is not None:
                                    if self.menu is not None and self.manualRun is False:
                                        self.menu.updateScenario(result[0], self.scenario_id, self.called)
                                        if result[0] == "FAILED":
                                            self.update_current_test_status("FAILED")

                                    self.write_line(row, wsx.max_column - 2, result[0])
                                    if result[0] != "No Auto Run":
                                        if self.manualRun is False:
                                            self.write_actual(row, wsx.max_column - 1, result[1])
                                        else:
                                            self.write_actual(row, wsx.max_column - 1, "x")
                                    if result[0] == "FAILED":
                                        color = "#FF0000"
                                    elif result[0] == "PASSED":
                                        color = "#90EE90"
                                    else:
                                        color = "#D3D3D3"

                                    self.log_write('')
                                    self.log_write(f'\tScenario result: {result[0]}.')
                                    # self.consoleprint("Scenario Result: %s\n%s" % (result[0], result[1]), color)

                                    self.consoleprint("Scenario Result: %s" % result[0], color)
                                    if len(result) > 2:
                                        if result[2] is not '':
                                            self.consoleprint("PASSED VALUES: %s" % result[2], "#90EE90")
                                    if len(result) > 3:
                                        if result[3] is not '':
                                            self.consoleprint("FAILED VALUES: %s" % result[3], "#FF0000")
                            except IndexError:
                                raise StopIteration
                # Wait for background video to end before running next scenario
                if self.video_end_time is not None:
                    self.consoleprint("Waiting for video to end...")
                    while time.time() < self.video_end_time:
                        self.sleep(1000)
                    self.consoleprint("...Done")
                    self.video_end_time = None

                    uncompressed_video_path = f'{self.result_path}/{self.currentsheet}/Scenario_{self.scenario_id}.avi'
                    compressed_video_path = uncompressed_video_path.replace(".avi", ".mp4")
                    self.compress_video(uncompressed_video_path, compressed_video_path)
                if self.manualRun and self.scenario_id != 0 and new_line is True:
                    title = "Scenario %d" % self.scenario_id
                    col = 0
                    reference_name = reference[column]
                    for label in headings:
                        col += 1
                        # run over table headers to find the expected results free text
                        if re.findall("^expected ", label.lower()) and reference_name is None:
                            try:
                                if openpyxl.__version__ <= "2.5.0-a3":
                                    val = str(ws.cell(None, row, col).value)
                                else:
                                    val = str(ws.cell(row, col).value)
                            except:
                                val = "Actual is like expected??"
                    answer = self.menu.popup(title, val, 3)
                    if answer == QMessageBox.Yes:
                        self.write_line(row, wsx.max_column - 2, "PASSED")
                        self.write_actual(row, wsx.max_column - 1, "As expected")
                        self.menu.updateScenario("PASSED", self.scenario_id)
                    else:
                        try:
                            actual = answer[1]
                        except:
                            actual = "N/A"
                        self.write_line(row, wsx.max_column - 2, "FAILED")
                        self.write_actual(row, wsx.max_column - 1, actual)
                        self.menu.updateScenario("FAILED", self.scenario_id)
                        self.update_current_test_status("FAILED")

                new_line = False
                if self.menu is not None:
                    self.progressSteps += 1
                    self.advance_progress_bar()
        except StopIteration:
            pass

    def find_hash_tag(self):
        ws = self.ws
        for row in range(1, ws.max_row):
            for col in range(1, ws.max_column):
                if openpyxl.__version__ <= "2.5.0-a3":
                    if '#' in str(ws.cell(None, row, col).value):
                        return [row, col]
                else:
                    if '#' in str(ws.cell(row, col).value):
                        return [row, col]
        return [0, 0]

    def find_scenario(self, scenario, column):
        ws = self.ws
        for row in range(1, ws.max_row):
            if openpyxl.__version__ <= "2.5.0-a3":
                if str(scenario) == str(ws.cell(None, row, column).value):
                    return row
            else:
                if str(scenario) == str(ws.cell(row, column).value):
                    return row
        return 0

    def advance_progress_bar(self):
        self.menu.progressBar.setValue(self.progressSteps + (self.steps - self.progressSteps) / 100)

    def sim_inject(self, label, value):
        """Injection to SIM label and value
        """
        try:
            unit_regex = re.search("(.*)(\[.*\])$", label)
            try:
                label = unit_regex.group(1)
                unit = unit_regex.group(2)
            except AttributeError:
                unit = "N/A"

            if type(value) == str:
                if value.lower() in self.actions:
                    value = self.actions[value.lower()]

            if re.findall("^noise_", str(value).lower()):
                noise_regex = re.search("^(?i)(noise_)(.*)\((.*)\)$", str(value))  # Split the string
                if noise_regex is None:
                    raise TestError(self.test_name, f"Cannot read {value}\nMake sure to follow the format!")

                noise_type = noise_regex.group(2).lower()  # Noise type
                new_value = noise_regex.group(3)  # Noise value
                x = new_value.split(',')  # Split values by ','
                if noise_type == "steps":
                    self.consoleprint(f"SimSTU.NoiseSteps({label},{unit},{x[0]},{x[1]},{x[2]},{x[3]},0,1)")

                    try:
                        self.SimEngineInterface.inject_noise_steps(label, unit, x[0], x[1], x[2], x[3])

                    except SimEngineInjectionError as error:
                        self.consoleprint(error.message, "red")
                        # raise TestError(self.testname, error.message)

                if noise_type == "sine":
                    self.consoleprint(f"SimSTU.NoiseSine({label},{unit},{x[0]},{x[1]},{x[2]},0,1)")

                    try:
                        self.SimEngineInterface.inject_noise_sine(label, unit, x[0], x[1], x[2])

                    except SimEngineInjectionError as error:
                        self.consoleprint(error.message, "red")
                        # raise TestError(self.testname, error.message)

                if noise_type == "constant":
                    self.consoleprint(f"SimSTU.NoiseConstant({label},{unit},{x[0]},{x[1]},1)")

                    try:
                        self.SimEngineInterface.inject_noise_constant(label, unit, x[0], int(x[1]), 1)

                    except SimEngineInjectionError as error:
                        self.consoleprint(error.message, "red")
                        # raise TestError(self.testname, error.message)

                if noise_type == "pulse":
                    self.consoleprint(f"SimSTU.NoisePulse({label},{unit},{x[0]},{x[1]},{x[2]},{x[3]},{x[4]})")
                    inject_value1 = x[0]
                    delay_time1 = int(x[1])
                    inject_value2 = x[2]
                    delay_time2 = int(x[3])
                    repeat_count = int(x[4])

                    try:
                        self.SimEngineInterface.inject_noise_pulse(label, unit, inject_value1, delay_time1,
                                                                   inject_value2, delay_time2, repeat_count)

                    except SimEngineInjectionError as error:
                        self.consoleprint(error.message, "red")
                        # raise TestError(self.testname, error.message)

                if noise_type == "manhattan":
                    repeat_count = int(x[x.__len__()-1])  # Repeat count shall be the last number
                    values = new_value[0:new_value.rindex(',')]  # Remove repeat count
                    self.consoleprint(f"SimSTU.NoiseManhattan({label},{unit},{values},{repeat_count})")
                    try:
                        self.SimEngineInterface.inject_noise_manhattan(label, unit, values, repeat_count)

                    except SimEngineInjectionError as error:
                        self.consoleprint(error.message, "red")
                        # raise TestError(self.testname, error.message)
            else:
                if type(value) == str:
                    if value.lower() == "true":
                        value = "True"
                    if value.lower() == "false":
                        value = "False"
                self.consoleprint("SIM: %s = %s" % (label, value))
                try:
                    self.SimEngineInterface.inject_value(label, unit, value)
                except SimEngineInjectionError as error:
                    self.consoleprint(error.message, "red")
                    # raise TestError(self.testname, error.message)

            self.sim_injections[label] = value

        except IndexError:
            raise TestError(self.test_name, "Incorrect usage of noise!!")

    def gdt_inject(self, dataitem, value, header):
        """Injection to GDT dataitem and value
        """
        try:
            ofp = 0
            if header == "GDT_PSP2":
                ofp = 1
            self.gdt_connect(ofp)

            injected = False
            if dataitem.lower() == "visor" and str(value).lower() == "true":
                # Shortcut word for VISOR_COEFFICIENTS injection to straighten the screen (Useful for symbology tests)
                af32PilotPoly3 = [[0.0] * 10, [0.0] * 10]
                af32PilotPoly3[0][4] = 1.0
                af32PilotPoly3[1][1] = 1.0
                for row in range(0, len(af32PilotPoly3)):
                    for col in range(0, len(af32PilotPoly3[0])):
                        injected = self.GDTInterfaces[ofp].inject_struct(
                            "VISOR_COEFFICIENTS", f"af32PilotPoly3[{row},{col}]", af32PilotPoly3[row][col])
                self.consoleprint(f"GDT_PSP{ofp+1}: VISOR_COEFFICIENTS = STRAIGHT")
                return injected

            # Inject data item
            if ".validity" in dataitem.lower():
                try:
                    value = Utilities.validity_or_number_to_bool(value)
                except ParameterError:
                    raise TestError(self.test_name, f"{dataitem}: \"{value}\" cannot be used for validity injection! value has to be Valid or Invalid.")

                dataitem_name = dataitem[:dataitem.index('.')]  # Remove .Validity from dataitem string
                method = "validity"
            elif ".override" in dataitem.lower():
                dataitem_name = dataitem[:dataitem.index('.')]  # Remove .override from dataitem string
                method = "override"
            else:
                dataitem_name = dataitem
                method = "value"

            for ItemType in self.GDTInterfaces[ofp].GDTInterfaceDataType:
                try:
                    if method == "validity":
                        is_injected = self.GDTInterfaces[ofp].inject_data_item_validity(ItemType, dataitem_name, value)
                    elif method == "override":
                        is_injected = self.GDTInterfaces[ofp].inject_data_item_override(ItemType, dataitem_name, value)
                    else:
                        is_injected = self.GDTInterfaces[ofp].inject_data_item(ItemType, dataitem, value)
                    if is_injected:
                        self.consoleprint(f"GDT_PSP{ofp+1}: %s = %s" % (dataitem, value))
                        self.gdt_injections[ofp][dataitem_name] = value, ItemType
                        injected = True
                        break
                except:
                    pass

            if not injected:
                raise TestError(self.test_name,
                                f'GDT error: Failed to inject value {value} into data item {dataitem}. Please check if the value or dataitem is valid.')

            self.sleep(SLEEP_TIME)
        except IndexError:
            raise TestError(self.test_name, "Invalid value!")

    def handle_action(self, *value):
        """Input: label/dataitem and value
        Returns array [PASS/FAIL, Actual]
        """
        function_name = value[0]
        step_counter = self.scenario_id
        if re.findall("^SLEEP|^DELAY", function_name.upper()):
            try:
                new_value = float(function_name[function_name.index('(') + 1:function_name.rindex(')')])
            except ValueError:
                raise TestError(self.test_name, "Wrong usage of delay, try delay(seconds)")
            self.consoleprint("delay (%d)" % new_value)
            self.sleep(new_value * 1000)
        if re.findall("^MSG", function_name.upper()):
            try:
                text = function_name[function_name.index('(') + 1:function_name.rindex(')')]
            except ValueError:
                text = ""
            if self.executed is False:
                self.menu.popup(Utilities.get_current_version(), text + "\n\nPress OK to continue test run", 1)
            else:
                if self.ci is False:
                    input("\n\n" + text + "\n\nPress OK to continue test run")
        if re.findall("^CALL_TEST", function_name.upper()):
            try:
                test = function_name[function_name.index('(') + 1:function_name.rindex(')')]
            except ValueError:
                test = ""
            self.consoleprint("Calling test: %s" % test)
            self.consoleprint("=" * 50)
            self.wb.create_sheet(test)
            call_test = TestClass(test, False, False, self.executed, True, self.menu, self.ci, False, True)
            call_test.caller = self

            call_test.run_test()

            # Write call_test status
            row = value[1]
            column = value[2]
            status = self.current_status
            self.write_line(row, column-2, status)
            self.write_actual(row, column-1, f"TEST_CALL:{test}")
            # reset vars after calltest was done
            if self.executed is False:
                self.menu.progressBar.setRange(0, self.steps)
        if re.findall("^SCREENSHOT", function_name.upper()):
            self.sleep(1000)
            screenshot_path = f'{self.result_path}/{self.currentsheet}/Scenario_{step_counter}.png'
            cmd = f'\"{screenshot_path}\"'
            if simulation:
                if self.host_env is True:
                    screenshot_request_status = self.socket_interface.request_screenshot()

                    if screenshot_request_status == PCSIMSocketInterface.PCSIMSocketInterfaceRequestStatus.SUCCESSFUL:
                        Image.open(Utilities.get_host_screenshot_path()).save(screenshot_path)
                        os.remove(Utilities.get_host_screenshot_path())

                    elif screenshot_request_status == PCSIMSocketInterface.PCSIMSocketInterfaceRequestStatus.ERROR_SCREENSHOT:
                        self.error_log('Failed saving screenshot. '
                                       'Please check if J:/ drive is mounted or if the simulation is running.')

                    else:
                        self.error_log('Failed saving screenshot. Unknown error.')
                else:
                    subprocess.Popen("v2u %s" % cmd, shell=True)

            self.sleep(1000)
            self.consoleprint("Success taking img -  %s" % cmd)

        if re.findall("^VIDEO", function_name.upper()) or re.findall("^BVIDEO", function_name.upper()):
            try:
                calc_time = float(function_name[function_name.index('(') + 1:function_name.rindex(')')])
            except ValueError:
                raise TestError(self.test_name, "Wrong usage of video, try video(seconds)")

            background_video = False
            if re.findall("^BVIDEO", function_name.upper()):
                background_video = True

            self.log_write("")
            self.log_write("\tTest Category: Video")

            uncompressed_video_path = f'{self.result_path}/{self.currentsheet}/Scenario_{str(step_counter)}.avi'
            compressed_video_path = uncompressed_video_path.replace(".avi", ".mp4")
            if self.host_env:
                self.consoleprint("Ignoring video recording (host environment).")
                self.log_write("\t\tIgnored video recording (host environment).")
                calc_time = float(function_name[function_name.index('(') + 1:function_name.rindex(')')])
                self.sleep(calc_time * 1000)
            else:
                calc_time_millisec = int(1000.0 * calc_time)
                self.consoleprint(f'Taking video: {uncompressed_video_path}')
                if simulation:
                    v2u_process = subprocess.Popen(
                        f'v2u -t {str(calc_time_millisec)} \"{uncompressed_video_path}\"', shell=True,
                        stdin=subprocess.PIPE, stderr=subprocess.PIPE, stdout=subprocess.PIPE)

                    if background_video:
                        self.sleep(300)
                        self.video_end_time = time.time() + calc_time
                    else:
                        if self.menu is not None:
                            self.sleep(calc_time_millisec)
                        v2u_process.wait()

                        self.compress_video(uncompressed_video_path, compressed_video_path)

                self.log_write(f'\t\tRecorded {str(calc_time)} seconds of video to file: {compressed_video_path}.')

        if re.findall("^GDT_DISCONNECT", function_name.upper()):
            for gdt_connection in self.GDTInterfaces:
                gdt_connection.disconnect()
                self.consoleprint(f"GDT: {gdt_connection.connection} Disconnected successfully.", "#3CB371")

    def handle_inject(self, *value):
        inject_type = value[0].upper()
        values = value[1]
        if inject_type == "GDT" or inject_type == "GDT_PSP2":
            dataitems = values.split(':')  # Split dataitems by ':'
            for data in dataitems:
                dataitem, inject_value = data.split('=')  # Split dataitem and value [dataitem,value]
                self.gdt_inject(dataitem, inject_value, inject_type)

        if inject_type == "SIM":
            sim_labels = values.split(':')  # Split labels by ':'
            for data in sim_labels:
                label, inject_value = data.split('=')  # Split label and value [label,value]
                self.sim_inject(label, inject_value)

    def handle_expected(self, *value):
        """Input: label/dataitem and value
        Returns array [PASS/FAIL, Actual]
        """
        function_name = value[0]
        step_counter = self.scenario_id

        frame_compare_regex = re.search('^(!?)FRAMECOMPARE', function_name.upper())
        if frame_compare_regex is not None:
            if frame_compare_regex.group(1) != '':
                is_absence = True
                is_absence_message = 'checking for absence'

            else:
                is_absence = False
                is_absence_message = 'checking for presence'

            try:
                new_value = function_name[function_name.index('(') + 1:function_name.rindex(')')]
            except ValueError:
                raise TestError(self.test_name, "Wrong usage of framecompare, please follow the format.")
            data = new_value.split(':')
            try:
                self.log_write('')
                self.log_write(f'\tTest Category: Golden Image Verification ({is_absence_message}).')

                golden_image_name = data[0]
                result_image_path = f'{self.result_path}/{self.currentsheet}/Scenario_{step_counter}.png'
                golden_image_path = f'{self.golden_images_path}/{golden_image_name}'
                output_image_path = f'{self.output_path}/{self.currentsheet}/Step {step_counter}'
                count = 0
                while os.path.isfile(output_image_path + ".png"):
                    if count == 0:
                        output_image_path = output_image_path + f"_{count}"
                    count += 1
                    output_image_path = output_image_path.replace("_" + str(count - 1), "_" + str(count))
                output_image_path += ".png"

                try:
                    image_comparator = ImageComparator(result_image_path, golden_image_path, output_image_path,
                                                       is_absence, self.log_write)

                except ImageComparatorError as e:
                    raise TestError(self.test_name, e.message)

                function_name = data[1]

                self.log_write(f'\t\tVerification method: {function_name.upper()}.')

                if function_name.upper() == 'SUB_IMAGE':
                    search_box_text = data[2]
                    search_box_regex = re.search("([0-9]+),([0-9]+),([0-9]+),([0-9]+)", search_box_text)

                    search_box_top_left = (int(search_box_regex.group(1)), int(search_box_regex.group(2)))
                    search_box_bottom_right = (int(search_box_regex.group(3)), int(search_box_regex.group(4)))

                    grab_box_text = data[3]
                    grab_box_regex = re.search("([0-9]+),([0-9]+),([0-9]+),([0-9]+)", grab_box_text)

                    grab_box_top_left = (int(grab_box_regex.group(1)), int(grab_box_regex.group(2)))
                    grab_box_bottom_right = (int(grab_box_regex.group(3)), int(grab_box_regex.group(4)))

                    tolerance = int(data[4])

                    try:
                        match_result = image_comparator.compare_sub_image(search_box_top_left, search_box_bottom_right,
                                                                          grab_box_top_left, grab_box_bottom_right, tolerance)

                    except ImageComparatorError as e:
                        raise TestError(self.test_name, e.message)

                elif function_name.upper() == 'PATTERN':
                    search_box_text = data[2]
                    box_regex = re.search("([0-9]+),([0-9]+),([0-9]+),([0-9]+)", search_box_text)

                    search_box_top_left = (int(box_regex.group(1)), int(box_regex.group(2)))
                    search_box_bottom_right = (int(box_regex.group(3)), int(box_regex.group(4)))

                    tolerance = int(data[3])

                    try:
                        match_result = image_comparator.compare_pattern(search_box_top_left, search_box_bottom_right,
                                                                        tolerance)

                    except ImageComparatorError as e:
                        raise TestError(self.test_name, e.message)

                elif function_name.upper() == 'PATTERN_RDP':
                    position_text = data[2]
                    position_regex = re.search("([a-zA-Z]+),([0-9.]+),([a-zA-Z]+),([0-9.]+)", position_text)

                    horizontal_indicator = position_regex.group(1)
                    horizontal_degrees = float(position_regex.group(2))
                    vertical_indicator = position_regex.group(3)
                    vertical_degrees = float(position_regex.group(4))

                    tolerance = int(data[3])

                    try:
                        match_result = image_comparator.compare_pattern_rdp(horizontal_indicator, horizontal_degrees,
                                                                            vertical_indicator, vertical_degrees,
                                                                            tolerance)

                    except ImageComparatorError as e:
                        raise TestError(self.test_name, e.message)

                else:
                    raise TestError(self.test_name, f'Invalid command: {function_name}.')

            except IndexError:
                msg = "Invalid Format! FrameCompare(source_file:element_type:position:tolerance)\n"
                if len(data) > 1:
                    if data[1].lower() == "sub_image":
                        msg = "Invalid Format! FrameCompare(source_file:element_type:search_box:grab_box:tolerance)\n"
                    if data[1].lower() == "pattern":
                        msg = "Invalid Format! FrameCompare(source_file:element_type:search_box:tolerance)\n"
                    if data[1].lower() == "pattern_rdp":
                        msg = "Invalid Format! FrameCompare(source_file:element_type:position:tolerance)\n"
                raise TestError(self.test_name, msg)

            if match_result:
                result = 'PASSED'
            else:
                result = 'FAILED'

            self.log_write(f'\t\tVerification result: {result}')
            actual_result_text = f"FRAMECOMPARE({data[0]}) = {result}"
            self.consoleprint(f'VISUAL_TESTING({result_image_path}): {actual_result_text}')
            return [result, actual_result_text]

        if re.findall("^VIDEOPROCESS", function_name.upper()):
            try:
                new_value = function_name[function_name.index('(') + 1:function_name.rindex(')')]
            except ValueError:
                raise TestError(self.test_name, "Wrong usage of framecompare, please follow the format.")
            data = new_value.split(':')
            if data[0].lower() == "flashing":
                try:
                    flash_count = int(data[1])
                    not_flashing_image = f"{self.golden_images_path}/{data[2]}"
                    flashing_image = f"{self.golden_images_path}/{data[3]}"

                    search_box = data[4].split(',')
                    left_corner = (int(search_box[0]), int(search_box[1]))
                    right_corner = (int(search_box[2]), int(search_box[3]))
                    tolerance = int(data[5])

                except IndexError:
                    raise TestError(self.test_name, "Wrong usage of videoprocess(flashing:count:steady_img:flashing_img:search_box:tolerance")

                result_video_path = f'{self.result_path}/{self.currentsheet}/Scenario_{step_counter}.mp4'
                output_video_path = f'{self.output_path}/{self.currentsheet}/Step {step_counter}.avi'

                try:
                    vid2 = VideoComparator(result_video_path, output_video_path, left_corner, right_corner,
                                           tolerance, 0.98)
                    result = vid2.flash_count(not_flashing_image, flashing_image, flash_count)
                except ImageComparatorError as error:
                    raise TestError(self.test_name, error.message)

                actual_result_text = f"VIDEOPROCESS = {result[1]}"
                self.consoleprint(f'VIDEO_PROCESS({result_video_path}): {actual_result_text}')
                return [result[0], actual_result_text]

        if re.findall("^SIM", function_name.upper()):
            self.log_write("")
            self.log_write("\tTest Category: SIMValue")
            try:
                # Remove the SIM() string
                new_value = function_name[function_name.index('(') + 1:function_name.rindex(')')]
            except ValueError:
                raise TestError(self.test_name, "Wrong usage of SIM data read, please follow the format:\n"
                                               "SIM(Label=value:Label2=value)")
            sim_labels = new_value.split(':')  # Split labels by ':'
            scenario_result = "PASSED"
            results_string = "SIMValues: "
            results_string_passed = ''
            results_string_failed = ''
            self.sleep(100)
            for data in sim_labels:
                current_result = "PASSED"
                label, expected_value = data.split('=')  # Split label and value [label,value]
                unit_regex = re.search("(.*)(\[.*\])$", label)
                try:
                    label = unit_regex.group(1)
                    unit = unit_regex.group(2)
                except AttributeError:
                    unit = "N/A"

                sim_value = self.SimEngineInterface.get_element_value(label, unit)
                # Result string to handle multiple values
                results_string += "%s = %s;" % (label, sim_value)
                if sim_value != expected_value:
                    scenario_result = "FAILED"
                    current_result = "FAILED"

                if current_result == "PASSED":
                    results_string_passed += "%s = %s;" % (label, sim_value)
                else:
                    results_string_failed += "%s = %s;" % (label, sim_value)

            self.log_write("\t\t%s" % results_string)
            self.log_write("\t\tSTAGE RESULT: " + scenario_result)
            return [scenario_result, results_string, results_string_passed, results_string_failed]

        if re.findall("^(GDT_STRUCT|GDT)(_PSP2)?(.*)", function_name.upper()):
            method_data = re.search("(GDT_STRUCT|GDT)(_PSP2)?\((.*)\)", function_name, re.IGNORECASE)
            try:
                method = method_data.group(1)
                ofp = 1 if method_data.group(2) is not None else 0
                data_items = method_data.group(3).split(':')  # Split dataitems by ':'
            except AttributeError:
                raise TestError(self.test_name, "Wrong usage of GDT struct read, please follow the format:\nGDT(dataitem=value)")

            self.gdt_connect(ofp)

            self.log_write("")
            self.log_write("\tTest Category: GDTValues")

            scenario_result = "PASSED"
            results_string = f"GDTVALUES: "
            results_string_passed = ''
            results_string_failed = ''

            if method == "GDT_STRUCT":
                struct_name = data_items[0]
                data_items = data_items[1:]
                results_string = f"GDTVALUES: ;{struct_name};"

            for data in data_items:
                current_result = "PASSED"

                data_result = re.search("(.*[^!<>])(!=|<=|>=|<|>|=)(.*)", data)  # Split dataitem, operator, value
                dataitem = data_result.group(1)
                operator_sign = data_result.group(2)
                expected_value = data_result.group(3)

                operator_dict = {
                    '!=': operator.ne,
                    '<': operator.lt,
                    '<=': operator.le,
                    '>': operator.gt,
                    '>=': operator.ge,
                    '=': operator.eq
                }
                if operator_sign not in operator_dict:
                    raise TestError(self.test_name, f"\'{operator_sign}\' is not a valid operator and cannot be used!\nValid operators are: !=, <= ,< ,>= ,>, =")

                self.sleep(1000)
                gdt_value = None

                if method == "GDT_STRUCT":
                    gdt_value = self.GDTInterfaces[ofp].read_struct(struct_name, dataitem)

                if method == "GDT_BITVIEW":
                    gdt_value = self.GDTInterfaces[ofp].read_buffer("Bit_Sequential_Counters", dataitem, "Value")

                if method == "GDT":
                    if ".validity" in dataitem.lower():
                        data_item_to_read = dataitem[:dataitem.index('.')]  # Remove .Validity from dataitem string
                        test_case = "validity"
                    else:
                        data_item_to_read = dataitem
                        test_case = "value"
                    buffer = None
                    for ItemType in self.GDTInterfaces[ofp].GDTInterfaceDataType:
                        try:
                            gdt_value = self.GDTInterfaces[ofp].read_data_item(ItemType, data_item_to_read)[test_case]
                            buffer = ItemType
                            if gdt_value is not None:
                                break
                        except:
                            pass

                if isinstance(gdt_value, bool):
                    try:
                        expected_value = Utilities.validity_or_number_to_bool(expected_value)
                    except ParameterError:
                        raise TestError(self.test_name, f"The expected value: \"{expected_value}\" cannot be used for validity comparison! value has to be Valid or Invalid.")

                if isinstance(gdt_value, float):
                    expected_value = float(expected_value)
                    if method == "GDT":
                        digits_to_show = 0
                        if buffer == GDTInterface.GDTInterfaceDataType.DOUBLE:
                            digits_to_show = 20
                        if buffer == GDTInterface.GDTInterfaceDataType.NUMBER:
                            digits_to_show = 7

                        if digits_to_show != 0:
                            digits = str(gdt_value).split(".")
                            digits_before_decimal = len(digits[0])
                            if "-" in digits[0]:
                                digits_before_decimal -= 1
                                expected_value = round(expected_value, digits_to_show - digits_before_decimal)

                if isinstance(gdt_value, int):
                    expected_value = int(expected_value)

                if operator_dict[operator_sign](gdt_value, expected_value) is False or isinstance(gdt_value, type(None)):
                    current_result = "FAILED"
                    scenario_result = "FAILED"

                if isinstance(gdt_value, bool):
                    # format boolean as valid/invalid
                    gdt_value = Utilities.bool_to_validity(gdt_value)

                if current_result == "PASSED":
                    results_string_passed += "%s = %s;" % (dataitem, gdt_value)
                else:
                    results_string_failed += "%s = %s;" % (dataitem, gdt_value)

                results_string += "%s = %s;" % (dataitem, gdt_value)

            self.log_write("\t\t%s" % results_string)
            self.log_write("\t\tSTAGE RESULT: " + scenario_result)
            return [scenario_result, results_string, results_string_passed, results_string_failed]

        if re.findall("N/A", function_name.upper()):
            self.log_write("\t\tSTAGE RESULT: N/A")
            return ["No Auto Run", ""]

    def write_line(self, row, col, value):

        if self.called is False:
            ws = self.ws
        else:
            ws = self.caller.wb[self.test_name]

        # actual = ws.cell(row=row, column=new_column,value=value)
        color = openpyxl.styles.colors.WHITE
        if "FAILED" in value:
            color = "FF0000"
        elif "PASSED" in value:
            color = "00b050"
        elif "No Auto Run" in value:
            color = "808080"
        # elif "Pass/Fail" or "Actual" in value: color = "bfbfbf" #4472c4
        elif "NO VALID LOG" in value:
            color = "FFFF8A"
        if "Empty" in value:
            value = None

        pass_fail = ws.cell(row=row, column=col, value=value)
        copy_style = ws.cell(row=row, column=col - 1)
        if copy_style.has_style:
            pass_fail._style = copy(copy_style._style)

        if color != openpyxl.styles.colors.WHITE:
            pass_fail.fill = openpyxl.styles.PatternFill('solid', color)

        # strExcelResultFile = '%s/%s/%s_result.xlsx'%(folder,self.testname,self.testname)
        # self.wb.save(strExcelResultFile)

    def write_actual(self, row, col, value):
        if self.called is False:
            ws = self.ws
        else:
            ws = self.caller.wb[self.test_name]
            # ws = self.ws
        # Actual column
        actual = ws.cell(row=row, column=col, value="")
        copy_style = ws.cell(row=row, column=col - 2)
        if copy_style.has_style:
            actual._style = copy(copy_style._style)

        if re.findall("^(!?)FRAMECOMPARE", value.upper()):
            actual.value = "Step %d.png" % self.scenario_id
            actual.hyperlink = "./Output/%s/Step %d.png" % (self.currentsheet, self.scenario_id)
        elif re.findall("^VIDEOPROCESS", value.upper()):
            actual.value = "Step %d.avi" % self.scenario_id
            actual.hyperlink = "./Output/%s/Step %d.avi" % (self.currentsheet, self.scenario_id)
        elif re.findall("^VIDEO", value.upper()):
            actual.value = "Scenario_%d.mp4" % self.scenario_id
            actual.hyperlink = "./Results/%s/Scenario_%d.mp4" % (self.currentsheet, self.scenario_id)
        elif re.findall("^SIMVALUES:", value.upper()):
            values = value[value.index(':') + 1:].replace(';', '\n').replace('\t', '')
            actual.value = "SIMValues: " + values
        elif re.findall("^GDTVALUES:", value.upper()):
            values = value[value.index(':') + 1:].replace(';', '\n').replace('\t', '')
            actual.value = "GDTVALUES: " + values
        elif re.findall("^TEST_CALL", value.upper()):
            values = value[value.index(':') + 1:]
            actual.hyperlink = f"#{values}!A1"
            actual.value = f"Sheet: \"{values}\""
        else:
            actual.value = value

            # strExcelResultFile = '%s/%s/%s_result.xlsx'%(folder,self.testname,self.testname)
            # self.wb.save(strExcelResultFile)

    def log_write(self, line):
        """
        with open(self.normal_log_file, 'a') as outfile:
            outfile.write(line + '\n')"""
        self.normal_logger.log_info(line)

    def error_log(self, string=''):
        self.error_logger.log_error(string + "\n\n\n")
        self.consoleprint(string, "red")

    def compress_video(self, uncompressed_video_path, compressed_video_path):
        ffmpeg_process = subprocess.Popen(f'ffmpeg -i \"{uncompressed_video_path}\" '
            f'-c:v libx264 -preset veryslow -crf 0 \"{compressed_video_path}\"', shell=True,
            stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=self.video_compression_log)
        ffmpeg_process.wait()

        os.remove(uncompressed_video_path)

    def gdt_connect(self, ofp):
        """
        This method checks if gdt is connected else try to connect
        """
        if self.GDTInterfaces[ofp].is_connected is False:
            self.GDTInterfaces[ofp].connect()
            if self.GDTInterfaces[ofp].is_connected:
                self.consoleprint(f"GDT: {self.GDTInterfaces[ofp].connection} Connected successfully.", "#3CB371")
            else:
                raise TestError(self.test_name, f'GDT {self.GDTInterfaces[ofp].connection} Connection failed. Please check the configurations.')

    def update_current_test_status(self, status):
        self.current_status = status
        if self.menu is not None:
            self.menu.updateCurrentStatus(status)
